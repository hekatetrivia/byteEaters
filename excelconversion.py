import openpyxl

class ExcelConversion(object):
    def __init__(self, dataset):
        workbook = openpyxl.load_workbook(dataset)
        sheet = workbook.active

        username = []
        for i in range(1,209):
            username.append(sheet.cell(i,1).value)
        with open('username', 'w') as f:
            f.write(str(username))

        labels = []
        for i in range(1,209):
            labels.append(sheet.cell(i,3).value)
        with open('labels', 'w') as f:
            #f.write(str(labels))
            for i in range(len(labels)):
                f.write(str(labels[i]))
                f.write('\n')


